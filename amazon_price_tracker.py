import re
import html
import time
import os
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Configuration ──────────────────────────────────────────────
SEARCH_QUERY = "notebook"
MAX_PAGES = 5
OUTPUT_FILE = "results.xlsx"
PAGE_LOAD_WAIT = 15
DEBUG = True


# ── Driver ─────────────────────────────────────────────────────
def create_driver():

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--lang=en-US")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    # _FOR USD ────────────────────────────────────────────────
    driver.get("https://www.amazon.com")
    driver.add_cookie({
        "name": "i18n-prefs",
        "value": "USD",
        "domain": ".amazon.com",
        "path": "/"
    })
    return driver


# ── URL builder ────────────────────────────────────────────────
# def build_search_url(query, page=1):
    # return f"https://www.amazon.com/s?k={query.strip().replace(' ', '+')}&page={page}"
def build_search_url(query, page=1):
    return f"https://www.amazon.com/s?k={query.strip().replace(' ', '+')}&page={page}&currency=USD"


# ── Page loader ────────────────────────────────────────────────
def load_page(driver, url):
    driver.get(url)
    for selector in [
        "[data-component-type='s-search-result']",
        "div.s-result-item[data-asin]",
        ".s-search-results .s-result-item",
    ]:
        try:
            WebDriverWait(driver, PAGE_LOAD_WAIT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            if DEBUG:
                print(f"  [debug] Loaded with: {selector}")
            return
        except TimeoutException:
            continue
    try:
        WebDriverWait(driver, PAGE_LOAD_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(3)
    except TimeoutException:
        print("  [!] Page failed to load.")


# ── Card finder ────────────────────────────────────────────────
def find_cards(driver):
    for selector in [
        "div.s-result-item[data-asin]",
        "[data-component-type='s-search-result']",
        ".s-result-item[data-asin]",
    ]:
        cards = [
            c for c in driver.find_elements(By.CSS_SELECTOR, selector)
            if c.get_attribute("data-asin")
        ]
        if cards:
            if DEBUG:
                print(f"  [debug] {len(cards)} cards with: {selector}")
            return cards
    return []


# ── Price cleaner ──────────────────────────────────────────────
def clean_price(raw):
    if not raw:
        return None
    decoded = html.unescape(raw)
    digits = re.sub(r"[^\d.]", "", decoded).rstrip(".")
    if not digits:
        return None
    try:
        return float(digits)
    except ValueError:
        return None


# ── Card parser ────────────────────────────────────────────────
def parse_card(card):
    asin = card.get_attribute("data-asin") or "N/A"

    title = (
        safe_text(card, "h2 .a-text-normal")
        or safe_text(card, "h2 a span")
        or safe_text(card, "h2 span")
        or "N/A"
    )
    if title == "N/A":
        return None
    # Shorten: cut at first dash, comma, or pipe
    for sep in [" - ", " | ", ", "]:
        if sep in title:
            title = title.split(sep)[0].strip()
            break
    if len(title) > 50:
        title = title[:50].rsplit(" ", 1)[0] + "..."

    raw_price = (
        safe_attr(card, ".a-price .a-offscreen", "innerHTML")
        or safe_attr(card, ".a-price .a-offscreen", "textContent")
        or safe_text(card, ".a-price-whole")
        or safe_text(card, ".a-color-price")
        or ""
    )
    price = clean_price(raw_price)

    rating = safe_attr(card, ".a-icon-alt", "innerHTML") or "N/A"
    if rating != "N/A":
        rating = html.unescape(rating).split(" ")[0]

    reviews = safe_text(card, ".a-size-base.s-underline-text") or "N/A"
    reviews = reviews.replace(",", "").strip()

    url = (
        safe_attr(card, "h2 a", "href")
        or safe_attr(card, "a.a-link-normal[href*='/dp/']", "href")
        or safe_attr(card, "a[href*='/dp/']", "href")
        or ""
    )
    if url and not url.startswith("http"):
        url = "https://www.amazon.com" + url.split("?")[0]
    if not url and asin != "N/A":
        url = f"https://www.amazon.com/dp/{asin}"

    return {
        "asin"   : asin,
        "title"  : title,
        "price"  : price,
        "rating" : rating,
        "reviews": reviews,
        "url"    : url or "N/A",
    }


# ── Helpers ────────────────────────────────────────────────────
def safe_text(parent, selector, default=""):
    try:
        el = parent.find_element(By.CSS_SELECTOR, selector)
        return (el.text or el.get_attribute("innerHTML") or "").strip()
    except NoSuchElementException:
        return default


def safe_attr(parent, selector, attr, default=""):
    try:
        el = parent.find_element(By.CSS_SELECTOR, selector)
        return (el.get_attribute(attr) or "").strip()
    except NoSuchElementException:
        return default


# ── Excel writer ───────────────────────────────────────────────
HEADERS = ["No", "ASIN", "Name", "Price", "Rating", "Reviews", "Link", "Scraped At"]

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(name="Arial", color="0563C1", underline="single", size=10)
NORMAL_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [5, 14, 60, 12, 10, 10, 12, 18]


def setup_sheet(ws, query):
    ws.title = "Results"

    # Header row
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def append_products(ws, products, start_row):
    for i, p in enumerate(products, start=start_row):
        row_num = i
        values = [
            row_num - 1,                                      # No
            p["asin"],                                        # ASIN
            p["title"],                                       # Name
            p["price"],                                       # Price (float)
            f"{p['rating']} ★" if p["rating"] != "N/A" else "N/A",  # Rating
            p["reviews"],                                     # Reviews
            None,                                             # Link (set below)
            datetime.now().strftime("%Y-%m-%d %H:%M"),        # Scraped At
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = CENTER if col != 3 else LEFT

        # Price formatting
        price_cell = ws.cell(row=row_num, column=4)
        price_cell.number_format = "$#,##0.00"

        # Clickable hyperlink in "Link" column
        link_cell = ws.cell(row=row_num, column=7)
        if p["url"] and p["url"] != "N/A":
            link_cell.value = "Open on Amazon"
            link_cell.hyperlink = p["url"]
            link_cell.font = LINK_FONT
            link_cell.alignment = CENTER
        else:
            link_cell.value = "N/A"

        # Zebra striping
        if row_num % 2 == 0:
            fill = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_num, column=col).fill = fill

        ws.row_dimensions[row_num].height = 18


def save_to_xlsx(products, start_row, filename=OUTPUT_FILE, query=""):
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        setup_sheet(ws, query)

    append_products(ws, products, start_row)
    wb.save(filename)


# ── Display ────────────────────────────────────────────────────
def print_separator(label=""):
    line = "─" * 65
    print(f"\n{line}\n  {label}\n{line}" if label else f"\n{line}")


def print_product(p, index):
    title = p["title"][:65] + "..." if len(p["title"]) > 65 else p["title"]
    price = f"${p['price']:,.2f}" if p["price"] else "N/A"
    print(f"""
  [{index}]
  Name   : {title}
  Price  : {price}
  Rating : {p['rating']}
  URL    : {p['url']}""")


# ── Main ───────────────────────────────────────────────────────
def run_scraper(query=SEARCH_QUERY, max_pages=MAX_PAGES):
    print_separator(f"Amazon Scraper  —  '{query}'")
    print(f"  Pages   : {max_pages}")
    print(f"  Output  : {OUTPUT_FILE}")

    driver = create_driver()
    print("\n  Browser ready (headless Chrome)")

    all_products = []

    try:
        for page in range(1, max_pages + 1):
            url = build_search_url(query, page)
            print_separator(f"Page {page} / {max_pages}")
            print(f"  {url}")

            load_page(driver, url)
            cards = find_cards(driver)

            if not cards:
                print("  [!] No cards found. Stopping.")
                break

            page_products = [p for p in (parse_card(c) for c in cards) if p]

            if not page_products:
                print("  [!] Cards found but could not parse any. Stopping.")
                break

            save_to_xlsx(page_products, start_row=len(all_products) + 2, query=query)
            all_products.extend(page_products)

            for i, p in enumerate(page_products, start=len(all_products) - len(page_products) + 1):
                print_product(p, i)

            print(f"\n  ✓ {len(page_products)} products from page {page}")

            if page < max_pages:
                time.sleep(3)

    except KeyboardInterrupt:
        print("\n\n  Stopped by user.")
    finally:
        driver.quit()
        print("\n  Browser closed.")

    print_separator()
    print(f"  Total : {len(all_products)} products")
    print(f"  Saved : {OUTPUT_FILE}  ← open in Excel, links are clickable!")
    print_separator()


if __name__ == "__main__":
    run_scraper()